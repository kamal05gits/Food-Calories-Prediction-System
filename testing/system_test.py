"""
SYSTEM TESTS — Food Calories Prediction System
===============================================
End-to-end tests of integrated components: the full prediction pipeline
(validate -> ML predict -> classify -> recommend -> store), prediction
accuracy, batch processing, and the FastAPI service as a whole.

Run:  python system_test.py
      (requires: pip install -r requirements.txt)
"""

import os
import shutil
import sys

import pandas as pd

import main
from fci import classify_food
from functions import calculate_calories

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
BATCH_PATH = os.path.join(BASE_DIR, "sample data input.csv")
MASTER_CSV = os.path.join(BASE_DIR, "food_prediction.csv")

RESULTS = []

CHICKEN = {"Food Name": "Grilled Chicken", "Serving Size (g)": "150", "Protein (g)": "28",
           "Carbohydrates (g)": "0", "Total Fat (g)": "3.6", "Dietary Fiber (g)": "0", "Sugars (g)": "0"}
FRIES = {"Food Name": "French Fries", "Serving Size (g)": "150", "Protein (g)": "3.4",
         "Carbohydrates (g)": "41", "Total Fat (g)": "15", "Dietary Fiber (g)": "3.4", "Sugars (g)": "0.6"}


def record(test_id, operation, expected, actual, passed):
    RESULTS.append({
        "Test ID": test_id,
        "Operation": operation,
        "Expected": expected,
        "Actual": actual,
        "Status": "Pass" if passed else "Fail",
    })


def report():
    print("\n### SYSTEM TEST RESULTS\n")
    print("| Test ID | Operation | Expected | Actual | Status |")
    print("|:---|:---|:---|:---|:---:|")
    for r in RESULTS:
        print(f"| {r['Test ID']} | {r['Operation']} | {r['Expected']} | {r['Actual']} | {r['Status']} |")
    passed = sum(1 for r in RESULTS if r["Status"] == "Pass")
    print(f"\n**{passed}/{len(RESULTS)} system tests passed.**\n")
    return passed == len(RESULTS)


def run():
    # ST-01 full pipeline consistency (validate -> predict -> classify)
    try:
        data, err = main.validate_inputs(dict(CHICKEN))
        rec, perr = main.predict_single(CHICKEN) if data is not None else (None, "validation failed")
        if rec:
            k = rec["Predicted Calories (kcal)"]
            cls, lvl = classify_food(k)
            ok = (perr is None and rec["Calorie Class"] == cls
                  and rec["Diet Caution Level"] == lvl)
            actual = f"{k} kcal -> {cls} / {lvl} (stored class={rec['Calorie Class']})"
        else:
            ok, actual = False, f"Failed: {perr}"
    except Exception as exc:
        ok, actual = False, f"Failed: {exc}"
    record("ST-01", "Full pipeline: validate -> predict -> classify (chicken)",
           "Stored class/level consistent with classify_food(prediction)", actual, ok)

    # ST-02 prediction accuracy vs Atwater reference
    try:
        rec, err = main.predict_single(CHICKEN)
        atwater = calculate_calories(28, 0, 3.6, 0, 150)          # 216.6 kcal
        k = rec["Predicted Calories (kcal)"]
        diff = abs(k - atwater)
        ok = err is None and diff <= 32
        actual = f"{k} kcal vs Atwater {atwater:.1f} kcal (diff {diff:.1f})"
    except Exception as exc:
        ok, actual = False, f"Failed: {exc}"
    record("ST-02", "ML prediction accuracy (chicken 150 g)",
           "Within +/-1 RMSE (32 kcal) of Atwater estimate", actual, ok)

    # ST-03 high-calorie food flows through the whole pipeline
    try:
        rec, err = main.predict_single(FRIES)
        k = rec["Predicted Calories (kcal)"]
        ok = (err is None and rec["Calorie Class"] == "High Calorie"
              and rec["Diet Caution Level"] == "High Caution"
              and "calorie-dense" in rec["Dietary Recommendation"])
        actual = (f"{k} kcal -> {rec['Calorie Class']} / {rec['Diet Caution Level']}, "
                  f"rec starts: {rec['Dietary Recommendation'][:40]}...")
    except Exception as exc:
        ok, actual = False, f"Failed: {exc}"
    record("ST-03", "Full pipeline for high-calorie food (fries 150 g)",
           "High Calorie / High Caution + calorie-dense recommendation", actual, ok)

    # ST-04 batch processing + master CSV storage (back up / restore log)
    try:
        backup = MASTER_CSV + ".bak_sys"
        if os.path.exists(MASTER_CSV):
            shutil.copy2(MASTER_CSV, backup)
        rows_before = len(pd.read_csv(MASTER_CSV)) if os.path.exists(MASTER_CSV) else 0
        out = main.predict_batch_csv(BATCH_PATH)
        stored = pd.read_csv(MASTER_CSV)
        ok = (len(out) == 7 and len(stored) == rows_before + 7
              and list(stored.columns) == main.RECORD_COLUMNS)
        actual = f"batch={len(out)} rows, master log {rows_before} -> {len(stored)}, columns ok"
    except Exception as exc:
        ok, actual = False, f"Failed: {exc}"
    finally:
        if os.path.exists(backup):
            shutil.move(backup, MASTER_CSV)
            backup = None
    record("ST-04", "Batch CSV -> predictions appended to master log",
           "7 rows processed, log grows by 7 with schema columns", actual, ok)

    # ST-05 XLSX storage integration
    tmp_xlsx = os.path.join(BASE_DIR, "_tmp_sys_xlsx.xlsx")
    try:
        if os.path.exists(tmp_xlsx):
            os.remove(tmp_xlsx)
        rec, _ = main.predict_single(CHICKEN)
        main.append_to_master_xlsx(pd.DataFrame([rec]), path=tmp_xlsx)
        main.append_to_master_xlsx(pd.DataFrame([rec]), path=tmp_xlsx)
        from openpyxl import load_workbook
        ws = load_workbook(tmp_xlsx).active
        ok = [c.value for c in ws[1]] == main.RECORD_COLUMNS and ws.max_row == 3
        actual = f"sheet '{ws.title}': 1 header + {ws.max_row - 1} rows"
    except Exception as exc:
        ok, actual = False, f"Failed: {exc}"
    finally:
        if os.path.exists(tmp_xlsx):
            os.remove(tmp_xlsx)
    record("ST-05", "XLSX workbook storage", "Workbook with header; one row per entry", actual, ok)

    # ST-06 FastAPI service end-to-end
    try:
        from fastapi.testclient import TestClient
        from api import app
        client = TestClient(app)
        code = client.get("/health").status_code
        p = client.post("/predict", json={
            "food_name": "French Fries", "serving_size": 150, "protein": 3.4,
            "carbohydrates": 41.0, "total_fat": 15.0, "dietary_fiber": 3.4, "sugars": 0.6,
        }).json()
        ok = (code == 200 and p.get("calorie_class") == "High Calorie"
              and p.get("diet_caution_level") == "High Caution"
              and 350 < p.get("predicted_calories", 0) <= 2000)
        actual = (f"health={code}, predict -> {p.get('predicted_calories')} kcal, "
                  f"{p.get('calorie_class')} / {p.get('diet_caution_level')}")
    except Exception as exc:
        ok, actual = False, f"Failed: {exc}"
    record("ST-06", "FastAPI service end-to-end (health + high-calorie prediction)",
           "200 health; /predict returns High Calorie / High Caution", actual, ok)

    return report()


if __name__ == "__main__":
    sys.exit(0 if run() else 1)

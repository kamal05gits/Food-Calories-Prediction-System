"""
ACCEPTANCE TESTS — Food Calories Prediction System
===================================================
User-story level acceptance criteria: does the system behave the way a
USER expects, from entry to stored record?

  AC-1  As a user, I enter food + nutrition details and get predicted
        calories for my serving.
  AC-2  As a user, I can understand the food's calorie classification
        and diet-caution level.
  AC-3  As a user, I receive an actionable dietary recommendation.
  AC-4  As a user, invalid input is handled gracefully (clear message,
        no crash).
  AC-5  As a user, I can process multiple foods from a CSV file.
  AC-6  As a user, every entry is stored (CSV master log + Excel file).

Run:  python acceptance_test.py
"""

import os
import shutil
import sys

import pandas as pd

import main
from functions import calculate_calories

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
BATCH_PATH = os.path.join(BASE_DIR, "sample data input.csv")
MASTER_CSV = os.path.join(BASE_DIR, "food_prediction.csv")

RESULTS = []

CHICKEN = {"Food Name": "Grilled Chicken", "Serving Size (g)": "150", "Protein (g)": "28",
           "Carbohydrates (g)": "0", "Total Fat (g)": "3.6", "Dietary Fiber (g)": "0", "Sugars (g)": "0"}
FRIES = {"Food Name": "French Fries", "Serving Size (g)": "150", "Protein (g)": "3.4",
         "Carbohydrates (g)": "41", "Total Fat (g)": "15", "Dietary Fiber (g)": "3.4", "Sugars (g)": "0.6"}


def record(test_id, story, expected, actual, passed):
    RESULTS.append({
        "Test ID": test_id,
        "User Story / Acceptance Criterion": story,
        "Expected": expected,
        "Actual": actual,
        "Status": "Pass" if passed else "Fail",
    })


def report():
    print("\n### ACCEPTANCE TEST RESULTS\n")
    print("| Test ID | User Story | Expected | Actual | Status |")
    print("|:---|:---|:---|:---|:---:|")
    for r in RESULTS:
        print(f"| {r['Test ID']} | {r['User Story / Acceptance Criterion']} "
              f"| {r['Expected']} | {r['Actual']} | {r['Status']} |")
    passed = sum(1 for r in RESULTS if r["Status"] == "Pass")
    print(f"\n**{passed}/{len(RESULTS)} acceptance criteria met.**\n")
    return passed == len(RESULTS)


def run():
    # AC-1 enter details -> get predicted calories
    try:
        rec, err = main.predict_single(CHICKEN)
        k = rec["Predicted Calories (kcal)"]
        atwater = calculate_calories(28, 0, 3.6, 0, 150)
        ok = err is None and 0.5 * atwater <= k <= 1.5 * atwater
        actual = f"Grilled Chicken 150 g -> {k} kcal (reasonable vs ~{atwater:.0f} kcal)"
    except Exception as exc:
        ok, actual = False, f"Failed: {exc}"
    record("AC-1", "Enter food details -> get predicted calories for the serving",
           "Prediction returned and in a sensible range", actual, ok)

    # AC-2 understand classification + caution level
    try:
        rec, err = main.predict_single(CHICKEN)
        ok = (err is None and rec["Calorie Class"] in ("Low Calorie", "Moderate Calorie")
              and rec["Diet Caution Level"] in ("Low Caution", "Medium Caution"))
        actual = f"{rec['Calorie Class']} / {rec['Diet Caution Level']} (not High for grilled chicken)"
    except Exception as exc:
        ok, actual = False, f"Failed: {exc}"
    record("AC-2", "See the food's calorie class and diet-caution level",
           "Levelling shown; lean protein NOT flagged High", actual, ok)

    # AC-3 actionable recommendation
    try:
        rec, err = main.predict_single(FRIES)
        r = rec["Dietary Recommendation"]
        ok = err is None and len(r) > 20 and ("limit" in r.lower() or "Priority" in r)
        actual = f"Recommendation: {r[:70]}..."
    except Exception as exc:
        ok, actual = False, f"Failed: {exc}"
    record("AC-3", "Receive an actionable dietary recommendation",
           "Non-empty, actionable text (e.g. 'limit portion size')", actual, ok)

    # AC-4 invalid input handled gracefully
    try:
        bad = {**CHICKEN, "Food Name": ""}
        data, err = main.validate_inputs(bad)
        ok = data is None and err is not None and len(err) > 5
        actual = f"Error message: '{err}' (no exception raised)"
    except Exception as exc:
        ok, actual = False, f"CRASHED: {exc}"
    record("AC-4", "Invalid input -> clear error, no crash",
           "User-facing error message returned cleanly", actual, ok)

    # AC-5 process multiple foods from CSV (back up / restore master log)
    try:
        backup = MASTER_CSV + ".bak_acc"
        if os.path.exists(MASTER_CSV):
            shutil.copy2(MASTER_CSV, backup)
        rows_before = len(pd.read_csv(MASTER_CSV)) if os.path.exists(MASTER_CSV) else 0
        out = main.predict_batch_csv(BATCH_PATH)
        ok = len(out) == 7 and "Predicted Calories (kcal)" in out.columns
        actual = f"All {len(out)} foods from CSV processed with calorie predictions"
    except Exception as exc:
        ok, actual = False, f"Failed: {exc}"
    finally:
        if os.path.exists(backup):
            shutil.move(backup, MASTER_CSV)
            backup = None
    record("AC-5", "Process multiple foods from a CSV file",
           "Every food in the file gets a prediction", actual, ok)

    # AC-6 every entry stored (CSV + XLSX)
    tmp_xlsx = os.path.join(BASE_DIR, "_tmp_acc_xlsx.xlsx")
    try:
        if os.path.exists(tmp_xlsx):
            os.remove(tmp_xlsx)
        rec, _ = main.predict_single(CHICKEN)
        main.append_to_master_csv(pd.DataFrame([rec])) if hasattr(main, "append_to_master_csv") else None
        df = pd.read_csv(MASTER_CSV)
        csv_ok = df.iloc[-1]["Food Name"] == "Grilled Chicken"
        main.append_to_master_xlsx(pd.DataFrame([rec]), path=tmp_xlsx)
        from openpyxl import load_workbook
        xlsx_ok = load_workbook(tmp_xlsx).active.max_row == 2
        ok = csv_ok and xlsx_ok
        actual = f"last CSV row='{df.iloc[-1]['Food Name']}', XLSX rows={load_workbook(tmp_xlsx).active.max_row - 1}"
    except Exception as exc:
        ok, actual = False, f"Failed: {exc}"
    finally:
        if os.path.exists(tmp_xlsx):
            os.remove(tmp_xlsx)
    record("AC-6", "Every entry is stored for record-keeping",
           "Entry present in CSV master log AND in the Excel file", actual, ok)

    return report()


if __name__ == "__main__":
    sys.exit(0 if run() else 1)

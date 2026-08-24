"""
REQUIREMENTS TESTS — Food Calories Prediction System
=====================================================
Verifies that each functional requirement stated in the README is actually
met by the deliverables (model file, dataset, API contract, batch support,
storage).

Requirements checked (from README):
    R-1  Trained ML model shipped (calorie_model.pkl) and loadable
    R-2  Dataset of 1,000 food records with all 6 features + calorie target
    R-3  API exposes /health (model load status)
    R-4  API /predict returns predicted calories + calorie class +
         diet-caution level + recommendation
    R-5  API enforces the documented input ranges
         (serving 1-1000 g, nutrients 0-100 g/100 g, fiber 0-40)
    R-6  Batch CSV prediction supported
    R-7  Prediction results stored in CSV master log AND separate XLSX

Run:  python requirements_test.py
      (requires: pip install -r requirements.txt)
"""

import os
import shutil
import sys

import joblib
import pandas as pd

import main
from functions import calculate_calories

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
MODEL_PATH = os.path.join(BASE_DIR, "calorie_model.pkl")
DATASET_PATH = os.path.join(BASE_DIR, "food_calorie_dataset_1000.csv")
BATCH_PATH = os.path.join(BASE_DIR, "sample data input.csv")
MASTER_CSV = os.path.join(BASE_DIR, "food_prediction.csv")

RESULTS = []


def record(test_id, requirement, expected, actual, passed):
    RESULTS.append({
        "Test ID": test_id,
        "Requirement": requirement,
        "Expected": expected,
        "Actual": actual,
        "Status": "Pass" if passed else "Fail",
    })


def report():
    print("\n### REQUIREMENTS TEST RESULTS\n")
    print("| Test ID | Requirement | Expected | Actual | Status |")
    print("|:---|:---|:---|:---|:---:|")
    for r in RESULTS:
        print(f"| {r['Test ID']} | {r['Requirement']} | {r['Expected']} | {r['Actual']} | {r['Status']} |")
    passed = sum(1 for r in RESULTS if r["Status"] == "Pass")
    print(f"\n**{passed}/{len(RESULTS)} requirements verified.**\n")
    return passed == len(RESULTS)


def api_client():
    from fastapi.testclient import TestClient
    from api import app
    return TestClient(app)


def run():
    # R-1 trained model artifact
    try:
        ok_file = os.path.exists(MODEL_PATH)
        model = joblib.load(MODEL_PATH)
        actual = f"file exists={ok_file}, loaded as {type(model).__name__}"
        ok = ok_file and model is not None
    except Exception as exc:
        actual, ok = f"Failed: {exc}", False
    record("R-1", "Trained model shipped and loadable", "calorie_model.pkl loads with joblib", actual, ok)

    # R-2 dataset
    try:
        df = pd.read_csv(DATASET_PATH)
        need = list(main.FEATURE_COLS) + ["Calories (kcal)"]
        ok = len(df) == 1000 and all(c in df.columns for c in need)
        actual = f"{len(df)} rows, columns ok={all(c in df.columns for c in need)}"
    except Exception as exc:
        actual, ok = f"Failed: {exc}", False
    record("R-2", "1,000-record dataset with 6 features + target", "1000 rows, all columns present", actual, ok)

    # R-3 API health endpoint
    try:
        h = api_client().get("/health").json()
        ok = h.get("model_loaded") is True
        actual = f"/health -> model_loaded={h.get('model_loaded')}"
    except Exception as exc:
        actual, ok = f"Failed: {exc}", False
    record("R-3", "API exposes /health", "model_loaded is true", actual, ok)

    # R-4 API prediction output schema
    try:
        p = api_client().post("/predict", json={
            "food_name": "Grilled Chicken", "serving_size": 150, "protein": 28,
            "carbohydrates": 0, "total_fat": 3.6, "dietary_fiber": 0, "sugars": 0,
        }).json()
        need = ["predicted_calories", "calorie_class", "diet_caution_level", "recommendation"]
        ok = all(k in p for k in need)
        actual = (f"predicted_calories={p.get('predicted_calories')}, "
                  f"class={p.get('calorie_class')}, caution={p.get('diet_caution_level')}, "
                  f"recommendation={'present' if p.get('recommendation') else 'MISSING'}")
    except Exception as exc:
        actual, ok = f"Failed: {exc}", False
    record("R-4", "/predict returns calories + class + caution level + recommendation",
           "All 4 output fields present", actual, ok)

    # R-5 input range enforcement
    client = api_client()
    bad_cases = [
        ("serving_size=0", {"serving_size": 0, "protein": 10, "carbohydrates": 10,
                            "total_fat": 5, "dietary_fiber": 2, "sugars": 5}),
        ("protein=150", {"serving_size": 100, "protein": 150, "carbohydrates": 10,
                         "total_fat": 5, "dietary_fiber": 2, "sugars": 5}),
        ("dietary_fiber=50", {"serving_size": 100, "protein": 10, "carbohydrates": 10,
                              "total_fat": 5, "dietary_fiber": 50, "sugars": 5}),
        ("missing field", {"serving_size": 100, "protein": 10}),
    ]
    ok = True
    details = []
    try:
        for label, body in bad_cases:
            code = client.post("/predict", json=body).status_code
            details.append(f"{label}->{code}")
            ok = ok and code == 422
        edge = client.post("/predict", json={"serving_size": 1000, "protein": 100,
                                             "carbohydrates": 100, "total_fat": 100,
                                             "dietary_fiber": 40, "sugars": 100}).status_code
        details.append(f"max-allowed-values->{edge}")
        ok = ok and edge == 200
    except Exception as exc:
        details, ok = [f"Failed: {exc}"], False
    record("R-5", "API enforces documented input ranges", "422 for out-of-range/missing, 200 at max allowed",
           ", ".join(details), ok)

    # R-6 batch CSV prediction (back up / restore the master log so the test is repeatable)
    try:
        backup = MASTER_CSV + ".bak_req"
        if os.path.exists(MASTER_CSV):
            shutil.copy2(MASTER_CSV, backup)
        rows_before = len(pd.read_csv(MASTER_CSV)) if os.path.exists(MASTER_CSV) else 0
        out = main.predict_batch_csv(BATCH_PATH)
        grew = os.path.exists(MASTER_CSV) and len(pd.read_csv(MASTER_CSV)) == rows_before + len(out)
        ok = len(out) == 7 and "Predicted Calories (kcal)" in out.columns and grew
        actual = f"{len(out)} records predicted, master log grew by {len(pd.read_csv(MASTER_CSV)) - rows_before}"
    except Exception as exc:
        actual, ok = f"Failed: {exc}", False
    finally:
        if os.path.exists(backup):
            shutil.move(backup, MASTER_CSV)
            backup = None
    record("R-6", "Batch CSV prediction supported", "sample CSV of 7 foods fully processed", actual, ok)

    # R-7 dual storage: CSV master (covered by R-6 growth) + XLSX workbook
    tmp_xlsx = os.path.join(BASE_DIR, "_tmp_req_xlsx.xlsx")
    try:
        if os.path.exists(tmp_xlsx):
            os.remove(tmp_xlsx)
        rec, _ = main.predict_single(VALID := {
            "Food Name": "Grilled Chicken", "Serving Size (g)": "150", "Protein (g)": "28",
            "Carbohydrates (g)": "0", "Total Fat (g)": "3.6", "Dietary Fiber (g)": "0", "Sugars (g)": "0",
        })
        main.append_to_master_xlsx(pd.DataFrame([rec]), path=tmp_xlsx)
        main.append_to_master_xlsx(pd.DataFrame([rec]), path=tmp_xlsx)
        from openpyxl import load_workbook
        ws = load_workbook(tmp_xlsx).active
        hdr_ok = [c.value for c in ws[1]] == main.RECORD_COLUMNS
        ok = hdr_ok and ws.max_row == 3
        actual = f"XLSX: header ok={hdr_ok}, {ws.max_row - 1} record rows appended"
    except Exception as exc:
        actual, ok = f"Failed: {exc}", False
    finally:
        if os.path.exists(tmp_xlsx):
            os.remove(tmp_xlsx)
    record("R-7", "Results stored in CSV master AND separate XLSX",
           "XLSX workbook created with header + appended rows (CSV growth checked in R-6)", actual, ok)

    return report()


if __name__ == "__main__":
    sys.exit(0 if run() else 1)

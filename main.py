import os
from datetime import datetime

import joblib
import numpy as np
import pandas as pd

from fci import classify_food, generate_recommendation

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
MODEL_PATH = os.path.join(BASE_DIR, "calorie_model.pkl")
MASTER_CSV_FILE = os.path.join(BASE_DIR, "food_prediction.csv")
MASTER_XLSX_FILE = os.path.join(BASE_DIR, "food_prediction.xlsx")

FEATURE_COLS = [
    "Serving Size (g)",
    "Protein (g)",
    "Carbohydrates (g)",
    "Total Fat (g)",
    "Dietary Fiber (g)",
    "Sugars (g)",
]
ID_COLS = ["Food Name"]
RECORD_COLUMNS = ID_COLS + FEATURE_COLS + [
    "Predicted Calories (kcal)",
    "Calorie Class",
    "Diet Caution Level",
    "Dietary Recommendation",
    "Timestamp",
]

NUMERIC_RANGES = {
    "Serving Size (g)": (1, 1000),
    "Protein (g)": (0, 100),
    "Carbohydrates (g)": (0, 100),
    "Total Fat (g)": (0, 100),
    "Dietary Fiber (g)": (0, 40),
    "Sugars (g)": (0, 100),
}

try:
    model = joblib.load(MODEL_PATH)
    MODEL_LOAD_ERROR = None
except Exception as exc:                       
    model = None
    MODEL_LOAD_ERROR = str(exc)


def validate_inputs(values):
    raw = {k: str(v).strip() for k, v in values.items()}

    required = ID_COLS + FEATURE_COLS
    if not all(raw.get(col) for col in required):
        return None, "Please fill in all input fields."

    data = {col: raw[col] for col in ID_COLS}
    for col in FEATURE_COLS:
        try:
            number = float(raw[col])
        except ValueError:
            return None, f"'{col}' must be a numerical value."
        low, high = NUMERIC_RANGES[col]
        if not (low <= number <= high):
            return None, f"'{col}' must fall between {low} and {high}."
        data[col] = number

    return data, None


def predict_record(data):
    if model is None:
        raise RuntimeError(
            f"Model could not be loaded: {MODEL_LOAD_ERROR}. "
            "Run train_model.py first."
        )

    features = np.array([[data[col] for col in FEATURE_COLS]], dtype=float)
    pred_calories = round(float(np.clip(model.predict(features)[0], 0, 2000)), 2)

    calorie_class, caution_level = classify_food(pred_calories)
    recommendation = generate_recommendation(
        caution_level,
        data["Serving Size (g)"], data["Protein (g)"], data["Carbohydrates (g)"],
        data["Total Fat (g)"], data["Dietary Fiber (g)"], data["Sugars (g)"],
    )

    record = dict(data)
    record.update({
        "Predicted Calories (kcal)": pred_calories,
        "Calorie Class": calorie_class,
        "Diet Caution Level": caution_level,
        "Dietary Recommendation": recommendation,
        "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    })
    return record


def predict_single(values):
    data, error = validate_inputs(values)
    if error:
        return None, error
    return predict_record(data), None


def predict_batch(df):
    if model is None:
        raise RuntimeError(
            f"Model could not be loaded: {MODEL_LOAD_ERROR}. "
            "Run train_model.py first."
        )

    missing = [c for c in FEATURE_COLS if c not in df.columns]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")

    df = df.copy()
    preds = np.clip(model.predict(df[FEATURE_COLS].values), 0, 2000)
    df["Predicted Calories (kcal)"] = np.round(preds, 2)
    df["Calorie Class"], df["Diet Caution Level"] = zip(
        *df["Predicted Calories (kcal)"].apply(classify_food)
    )
    df["Dietary Recommendation"] = df.apply(
        lambda r: generate_recommendation(
            r["Diet Caution Level"], r["Serving Size (g)"], r["Protein (g)"],
            r["Carbohydrates (g)"], r["Total Fat (g)"],
            r["Dietary Fiber (g)"], r["Sugars (g)"],
        ),
        axis=1,
    )
    df["Timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return df


def predict_batch_csv(file_path):
    df = pd.read_csv(file_path)
    df = predict_batch(df)
    save_cols = [c for c in RECORD_COLUMNS if c in df.columns]
    append_to_master_csv(df[save_cols])
    append_to_master_xlsx(df[save_cols])
    return df[save_cols]


def append_to_master_csv(df_to_add):
    if not os.path.exists(MASTER_CSV_FILE):
        df_to_add.to_csv(MASTER_CSV_FILE, mode="w", header=True, index=False)
    else:
        df_to_add.to_csv(MASTER_CSV_FILE, mode="a", header=False, index=False)


def append_to_master_xlsx(df_to_add, path=None):
    from openpyxl import Workbook, load_workbook

    path = path or MASTER_XLSX_FILE
    columns = [c for c in RECORD_COLUMNS if c in df_to_add.columns] \
        or list(df_to_add.columns)

    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Food Predictions"
        ws.append(columns)

    for _, row in df_to_add.iterrows():
        ws.append([row.get(c, "") for c in columns])
    wb.save(path)


def launch_app():
    import tkinter as tk
    from tkinter import filedialog, messagebox

    root = tk.Tk()
    root.title("Food Calorie Prediction System")
    root.geometry("1100x780")
    root.resizable(True, True)
    root.configure(bg="#f4f6f9")

    main_frame = tk.Frame(root, padx=25, pady=20, bg="#f4f6f9")
    main_frame.pack(fill="both", expand=True)

    tk.Label(main_frame, text="FOOD CALORIE PREDICTION SYSTEM",
             font=("Helvetica", 18, "bold"), bg="#f4f6f9",
             fg="#b23a1d").pack(pady=(0, 15))

    form_frame = tk.Frame(main_frame, bg="#f4f6f9")
    form_frame.pack(fill="x", pady=10)

    f_left = tk.LabelFrame(form_frame, text="Food Information",
                           font=("Helvetica", 11, "bold"),
                           padx=15, pady=10, bg="#f4f6f9")
    f_left.pack(side="left", fill="both", expand=True, padx=(0, 10))

    tk.Label(f_left, text="Food Name:", bg="#f4f6f9").grid(row=0, column=0, sticky="w", pady=4)
    entry_name = tk.Entry(f_left, width=22)
    entry_name.grid(row=0, column=1, pady=4)

    f_right = tk.LabelFrame(form_frame, text="Nutritional Information (per 100 g)",
                            font=("Helvetica", 11, "bold"),
                            padx=15, pady=10, bg="#f4f6f9")
    f_right.pack(side="left", fill="both", expand=True, padx=(10, 0))

    field_labels = [
        ("Serving Size (g):", "Serving Size (g)"),
        ("Protein (g):", "Protein (g)"),
        ("Carbohydrates (g):", "Carbohydrates (g)"),
        ("Total Fat (g):", "Total Fat (g)"),
        ("Dietary Fiber (g):", "Dietary Fiber (g)"),
        ("Sugars (g):", "Sugars (g)"),
    ]
    entries = {}
    for i, (label_text, key) in enumerate(field_labels):
        tk.Label(f_right, text=label_text, bg="#f4f6f9").grid(
            row=i, column=0, sticky="w", pady=2)
        e = tk.Entry(f_right, width=20)
        e.grid(row=i, column=1, pady=2)
        entries[key] = e

    result_frame = tk.LabelFrame(main_frame, text="Prediction Results",
                                 font=("Helvetica", 12, "bold"),
                                 padx=15, pady=10, bg="#ffffff")
    result_frame.pack(fill="both", padx=5, pady=10)

    lbl_pred = tk.Label(result_frame, text="Predicted Calories: --",
                        font=("Helvetica", 11), bg="#ffffff")
    lbl_pred.pack(anchor="w", pady=2)
    lbl_class = tk.Label(result_frame, text="Calorie Class: --",
                         font=("Helvetica", 11), bg="#ffffff")
    lbl_class.pack(anchor="w", pady=2)
    lbl_caution = tk.Label(result_frame, text="Diet Caution Level: --",
                           font=("Helvetica", 11), bg="#ffffff")
    lbl_caution.pack(anchor="w", pady=2)
    lbl_rec = tk.Label(result_frame, text="Recommendation: --",
                       font=("Helvetica", 10, "italic"), fg="#0d6efd",
                       bg="#ffffff", wraplength=950, justify="left")
    lbl_rec.pack(anchor="w", pady=4)

    def on_predict_single():
        if model is None:
            messagebox.showerror(
                "Model Missing",
                "Model file not found. Run train_model.py first.")
            return
        values = {"Food Name": entry_name.get()}
        values.update({key: ent.get() for key, ent in entries.items()})

        record, error = predict_single(values)
        if error:
            messagebox.showerror("Invalid Input", error)
            return

        lbl_pred.config(
            text=f"Predicted Calories: {record['Predicted Calories (kcal)']} kcal")
        lbl_class.config(text=f"Calorie Class: {record['Calorie Class']}")
        lbl_caution.config(text=f"Diet Caution Level: {record['Diet Caution Level']}")
        lbl_rec.config(text=f"Recommendation: {record['Dietary Recommendation']}")

        append_to_master_csv(pd.DataFrame([record]))
        append_to_master_xlsx(pd.DataFrame([record]))
        messagebox.showinfo(
            "Success", f"Prediction calculated and saved to "
                       f"'{os.path.basename(MASTER_CSV_FILE)}' and "
                       f"'{os.path.basename(MASTER_XLSX_FILE)}'.")

    def on_predict_batch_csv():
        if model is None:
            messagebox.showerror("Error", "Model file not loaded.")
            return
        file_path = filedialog.askopenfilename(
            title="Select CSV Dataset", filetypes=[("CSV Files", "*.csv")])
        if not file_path:
            return
        try:
            out = predict_batch_csv(file_path)
            messagebox.showinfo(
                "Success", f"Processed {len(out)} records into "
                           f"'{os.path.basename(MASTER_CSV_FILE)}' and "
                           f"'{os.path.basename(MASTER_XLSX_FILE)}'.")
        except Exception as exc:
            messagebox.showerror("Processing Failed", str(exc))

    def on_clear():
        clear_fields([entry_name] + list(entries.values()),
                     [lbl_pred, lbl_class, lbl_caution, lbl_rec])
        entry_name.focus()

    # ---- 9.3 Action Section ----
    btn_frame = tk.Frame(main_frame, bg="#f4f6f9")
    btn_frame.pack(pady=15)

    tk.Button(btn_frame, text="Predict Calories", command=on_predict_single,
              bg="#b23a1d", fg="white", width=16,
              font=("Helvetica", 10, "bold")).pack(side="left", padx=5)
    tk.Button(btn_frame, text="Batch Predict CSV", command=on_predict_batch_csv,
              bg="#008272", fg="white", width=16,
              font=("Helvetica", 10, "bold")).pack(side="left", padx=5)
    tk.Button(btn_frame, text="Clear", command=on_clear,
              width=10).pack(side="left", padx=5)
    tk.Button(btn_frame, text="Exit", command=root.destroy,
              bg="#d9534f", fg="white", width=10).pack(side="left", padx=5)

    root.mainloop()


def clear_fields(entry_widgets, label_widgets):
    import tkinter as tk
    for entry in entry_widgets:
        entry.delete(0, tk.END)
    defaults = ["Predicted Calories: --", "Calorie Class: --",
                "Diet Caution Level: --", "Recommendation: --"]
    for label, text in zip(label_widgets, defaults):
        label.config(text=text)


if __name__ == "__main__":
    launch_app()

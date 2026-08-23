import os

import joblib
import pandas as pd

import main
from fci import classify_food, generate_recommendation
from functions import calculate_calories, calculate_calorie_level

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
MODEL_PATH = os.path.join(BASE_DIR, "calorie_model.pkl")
DATASET_PATH = os.path.join(BASE_DIR, "food_calorie_dataset_1000.csv")
BATCH_PATH = os.path.join(BASE_DIR, "sample_batch_input.csv")
MASTER_CSV = os.path.join(BASE_DIR, "food_prediction.csv")

RESULTS = []


def record(test_id, module, operation, expected, actual, passed):
    RESULTS.append({
        "Test ID": test_id,
        "Module": module,
        "Test Operation": operation,
        "Expected Result": expected,
        "Actual Result": actual,
        "Status": "Pass" if passed else "Fail",
    })
    return passed


def run_tests():
    # TO-01  Model file loads ----------------------------------------
    try:
        m = joblib.load(MODEL_PATH)
        actual = "Model loaded successfully"
        ok = m is not None
    except Exception as exc:
        actual, ok = f"Load failed: {exc}", False
    record("TO-01", "train_model.py / calorie_model.pkl",
           "Load trained model file with joblib",
           "Model loads without error", actual, ok)

    # TO-02  Dataset integrity ---------------------------------------
    try:
        df = pd.read_csv(DATASET_PATH)
        ok = len(df) == 1000 and all(
            c in df.columns for c in main.FEATURE_COLS + ["Calories (kcal)"])
        actual = f"{len(df)} records, all required columns present"
    except Exception as exc:
        actual, ok = f"Dataset error: {exc}", False
    record("TO-02", "food_calorie_dataset_1000.csv",
           "Check dataset record count and required columns",
           "1,000 records with all feature + target columns", actual, ok)

    # TO-03  validate_inputs - valid data -----------------------------
    valid = {
        "Food Name": "Grilled Chicken",
        "Serving Size (g)": "150", "Protein (g)": "28",
        "Carbohydrates (g)": "0", "Total Fat (g)": "3.6",
        "Dietary Fiber (g)": "0", "Sugars (g)": "0",
    }
    data, err = main.validate_inputs(valid)
    record("TO-03", "main.validate_inputs()",
           "Validate a complete, in-range record",
           "Returns parsed data, no error",
           f"Error: {err}" if err else "Record accepted, no error",
           data is not None and err is None)

    # TO-04  validate_inputs - empty fields ---------------------------
    _, err = main.validate_inputs({**valid, "Food Name": ""})
    record("TO-04", "main.validate_inputs()",
           "Submit form with an empty mandatory field",
           "Validation error raised",
           f"Error: {err}",
           err is not None and "all input fields" in err)

    # TO-05  validate_inputs - non-numeric value ----------------------
    _, err = main.validate_inputs({**valid, "Protein (g)": "abc"})
    record("TO-05", "main.validate_inputs()",
           "Enter non-numeric value in a nutrient field",
           "'must be a numerical value' error",
           f"Error: {err}",
           err is not None and "numerical" in err)

    # TO-06  validate_inputs - range error ----------------------------
    _, err = main.validate_inputs({**valid, "Total Fat (g)": "150"})
    record("TO-06", "main.validate_inputs()",
           "Enter out-of-range nutrient value (fat = 150 g/100 g)",
           "Range error raised",
           f"Error: {err}",
           err is not None and "between 0 and 100" in err)

    # TO-07  validate_inputs - serving size range ----------------------
    _, err = main.validate_inputs({**valid, "Serving Size (g)": "0"})
    record("TO-07", "main.validate_inputs()",
           "Enter invalid serving size (0 g)",
           "Range error raised",
           f"Error: {err}",
           err is not None and "between 1 and 1000" in err)

    # TO-08  Single ML prediction --------------------------------------
    rec, err = main.predict_single(valid)
    atwater = calculate_calories(28, 0, 3.6, 0, 150)      # 216.6 kcal
    ok = rec is not None and abs(rec["Predicted Calories (kcal)"] - atwater) <= 32
    actual = (f"{rec['Predicted Calories (kcal)']} kcal "
              f"(Atwater estimate {atwater:.1f} kcal, diff "
              f"{abs(rec['Predicted Calories (kcal)'] - atwater):.1f} kcal)") if rec else f"Error: {err}"
    record("TO-08", "main.predict_single() + model",
           "Predict calories for grilled chicken (150 g)",
           "Prediction within +/-1 RMSE (32 kcal) of Atwater estimate (~216.6 kcal)",
           actual, ok)

    # TO-09 / 10 / 11  Calorie classification boundaries ---------------
    c_low = classify_food(120)
    c_mod = classify_food(250)
    c_high = classify_food(480)
    record("TO-09", "fci.classify_food()",
           "Classify 120 kcal serving",
           "'Low Calorie' / 'Low Caution'", f"{c_low}",
           c_low == ("Low Calorie", "Low Caution"))
    record("TO-10", "fci.classify_food()",
           "Classify 250 kcal serving",
           "'Moderate Calorie' / 'Medium Caution'", f"{c_mod}",
           c_mod == ("Moderate Calorie", "Medium Caution"))
    record("TO-11", "fci.classify_food()",
           "Classify 480 kcal serving",
           "'High Calorie' / 'High Caution'", f"{c_high}",
           c_high == ("High Calorie", "High Caution"))

    # TO-12  Recommendation flags high fat + sugar ----------------------
    rec_txt = generate_recommendation("High Caution", 90, 4.2, 52, 22.5, 1.8, 38.0)
    ok = "smaller portion" in rec_txt and "sugar" in rec_txt and rec_txt.startswith("Priority")
    record("TO-12", "fci.generate_recommendation()",
           "Recommendation for high-fat, high-sugar dessert",
           "Priority actions mentioning portion/fat and sugar",
           rec_txt, ok)

    # TO-13  Recommendation for balanced low-caution food ---------------
    rec_txt = generate_recommendation("Low Caution", 180, 0.3, 13.8, 0.2, 2.4, 10.4)
    ok = "Enjoy freely" in rec_txt or "balanced diet" in rec_txt
    record("TO-13", "fci.generate_recommendation()",
           "Recommendation for a low-caution food (apple)",
           "Positive 'enjoy freely / fits diet' message",
           rec_txt, ok)

    # TO-14  Batch CSV prediction ---------------------------------------
    existed_before = os.path.exists(MASTER_CSV)
    rows_before = len(pd.read_csv(MASTER_CSV)) if existed_before else 0
    try:
        out = main.predict_batch_csv(BATCH_PATH)
        ok = len(out) == 7 and "Predicted Calories (kcal)" in out.columns
        actual = f"{len(out)} records processed and appended"
    except Exception as exc:
        actual, ok = f"Batch failed: {exc}", False
    record("TO-14", "main.predict_batch_csv()",
           "Batch-predict sample_batch_input.csv (7 foods)",
           "7 records predicted with output columns", actual, ok)

    # TO-15  Prediction record storage ----------------------------------
    try:
        stored = pd.read_csv(MASTER_CSV)
        grew = len(stored) == rows_before + 7
        cols_ok = list(stored.columns) == main.RECORD_COLUMNS
        actual = f"food_prediction.csv has {len(stored)} rows with correct columns"
        ok = grew and cols_ok
    except Exception as exc:
        actual, ok = f"Storage check failed: {exc}", False
    record("TO-15", "main.append_to_master_csv()",
           "Verify batch results appended to food_prediction.csv",
           "Row count grows by 7, columns match record schema", actual, ok)

    # TO-16  Rule-based fallback (Day 1 prototype) -----------------------
    # 10 g protein, 10 g carbs, 5 g fat, 5 g fiber per 100 g, 200 g serving:
    # (10*4 + 10*4 + 5*9 + 5*2) * 2 = 135 * 2 = 270 kcal exactly.
    kcal = calculate_calories(10, 10, 5, 5, 200)
    level = calculate_calorie_level(kcal)
    ok = abs(kcal - 270.0) < 0.01 and level == "Moderate Calorie"
    record("TO-16", "functions.py (rule-based prototype)",
           "Atwater estimate for 200 g serving (135 kcal/100 g food)",
           "Exactly 270.0 kcal -> Moderate Calorie",
           f"{kcal:.2f} kcal -> {level}", ok)

    # TO-17  FastAPI endpoints -------------------------------------------
    try:
        from fastapi.testclient import TestClient
        from api import app as api_app
        client = TestClient(api_app)
        h = client.get("/health").json()
        p = client.post("/predict", json={
            "food_name": "Grilled Chicken",
            "serving_size": 150, "protein": 28, "carbohydrates": 0,
            "total_fat": 3.6, "dietary_fiber": 0, "sugars": 0}).json()
        ok = h.get("model_loaded") and "predicted_calories" in p
        actual = (f"/health model_loaded={h.get('model_loaded')}, "
                  f"/predict -> {p.get('predicted_calories')} kcal "
                  f"({p.get('calorie_class')})")
    except Exception as exc:
        actual, ok = f"API test failed: {exc}", False
    record("TO-17", "api.py (FastAPI)",
           "Call /health and /predict endpoints",
           "Health OK and valid prediction JSON returned", actual, ok)

    # TO-18  XLSX (Excel) storage of user entries --------------------------
    tmp_xlsx = os.path.join(BASE_DIR, "_tmp_to18.xlsx")
    try:
        if os.path.exists(tmp_xlsx):
            os.remove(tmp_xlsx)
        rec, _ = main.predict_single(valid)
        main.append_to_master_xlsx(pd.DataFrame([rec]), path=tmp_xlsx)
        main.append_to_master_xlsx(pd.DataFrame([rec]), path=tmp_xlsx)
        from openpyxl import load_workbook
        wb = load_workbook(tmp_xlsx)
        ws = wb.active
        hdr_ok = [c.value for c in ws[1]] == main.RECORD_COLUMNS
        ok = hdr_ok and ws.max_row == 3
        actual = (f"workbook sheet '{ws.title}': 1 header + "
                  f"{ws.max_row - 1} record rows appended")
    except Exception as exc:
        actual, ok = f"XLSX test failed: {exc}", False
    finally:
        if os.path.exists(tmp_xlsx):
            os.remove(tmp_xlsx)
    record("TO-18", "main.append_to_master_xlsx()",
           "Store user entry in the separate food_prediction.xlsx",
           "Workbook created with header; rows appended on each entry",
           actual, ok)

    # ----------------------------------------------------------------- #
    print("\n### Test Operation Results\n")
    print("| Test ID | Module / Function | Test Operation | Expected Result | Actual Result | Status |")
    print("|:--------|:------------------|:---------------|:----------------|:--------------|:------:|")
    for r in RESULTS:
        print(f"| {r['Test ID']} | {r['Module']} | {r['Test Operation']} "
              f"| {r['Expected Result']} | {r['Actual Result']} | {r['Status']} |")

    passed = sum(1 for r in RESULTS if r["Status"] == "Pass")
    print(f"\n**{passed}/{len(RESULTS)} test operations passed.**\n")


if __name__ == "__main__":
    run_tests()
